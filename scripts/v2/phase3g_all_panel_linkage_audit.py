"""Build the diagnostic-only Phase-3G all-panel identity/linkage audit.

The program reads raw genotype identifiers as strings, keeps panel namespaces
separate, and never mutates Phase-3 or certified-v1 artifacts.  Large matrices
are profiled from their headers or hash-bound prior line counts; genotype calls
are not materialized and no kernel is constructed.
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import re
from typing import Iterable

import duckdb
import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from scripts.v2.phase3g_identifier_semantics import (
    panel_sample_key,
    parse_identifier,
    resolve_gid_candidates,
)
from scripts.v2.phase3g_r2_semantics import (
    PARSER_VERSION as R2_PARSER_VERSION,
    canonical_80k_pav_axes,
    parse_hibap_sources,
)


VERSION = "phase3g_all_panel_genotype_linkage_audit_v2"
SELECTED_TRAITS = {
    "1000_GRAIN_WEIGHT",
    "ABOVE_GROUND_BIOMASS",
    "DAYS_TO_HEADING",
    "DAYS_TO_MATURITY",
    "GRAIN_YIELD",
    "PLANT_HEIGHT",
    "TEST_WEIGHT",
}

EXPECTED_ORIGINAL = {
    "stage1_gids": 16_579,
    "hmp": 5_253,
    "dartag": 1_931,
    "hibap": 96,
    "dartag_or_hibap": 2_027,
    "hmp_and_dartag_or_hibap": 1_382,
    "original_union": 5_898,
    "outside_original_union": 10_681,
    "selected_gids": 16_557,
    "selected_hmp": 5_253,
    "selected_original_union": 5_898,
    "selected_rows": 3_193_677,
    "selected_rows_original_union": 1_324_217,
}


def clean(value: object) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return str(value).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_path(value: object) -> str:
    return clean(value).replace("\\", "/")


def authoritative_gid(value: object, context: str = "authoritative_gid_column", *, excel: bool = False) -> tuple[str, str]:
    decision = parse_identifier(value, context=context, excel_derived=excel)
    return decision.canonical_gid_candidate, decision.normalization_rule


def delimited_rows(path: Path, delimiter: str = ",", limit: int | None = None) -> Iterable[list[str]]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader):
            if limit is not None and index >= limit:
                break
            yield row


def first_noncomment_rows(path: Path, delimiter: str, count: int) -> list[list[str]]:
    values: list[list[str]] = []
    for row in delimited_rows(path, delimiter=delimiter):
        if len(row) == 1 and clean(row[0]).startswith("#"):
            continue
        values.append(row)
        if len(values) == count:
            break
    return values


def xlsx_header(path: Path, sheet_name: str, required: set[str], max_rows: int = 40) -> tuple[int, dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    try:
        for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            names = {clean(value).upper(): index for index, value in enumerate(values) if clean(value)}
            if required.issubset(names):
                return row_number, names
            if row_number >= max_rows:
                break
    finally:
        workbook.close()
    raise RuntimeError(f"No header with {sorted(required)} in {path}::{sheet_name}")


def panel_id_for_file(relative_path: str) -> tuple[str, str, str]:
    rel = normalize_path(relative_path)
    lower = rel.lower()
    dataset = rel.split("/", 1)[0]
    if dataset.startswith("57IBWSN"):
        return "mas_57ibwsn_42sawsn_35hrwsn", "KASP/STS/SSR", "gene_based_mas"
    if dataset.startswith("58IBWSN"):
        return "mas_58ibwsn_43sawsn", "KASP/STS/SSR", "gene_based_mas"
    if dataset == "80k":
        population = next((name for name in ("hexaploid", "tetraploid", "wheat_recall", "wild_relative") if name in lower), "collection")
        return f"dartseq80k_{population}", "DArTseq 80K", "DArTseq"
    if dataset.startswith("DArTseq-derived"):
        return "mexican_landrace_dartseq", "DArTseq", "DArTseq"
    if dataset == "GBS":
        parent = rel.split("/")[1].lower() if "/" in rel else ""
        for number in range(13, 19):
            if parent.startswith(str(number)):
                return f"gbs_{number}sawyt", "GBS", "GBS"
        return "mas_45ibwsn", "MAS", "gene_based_mas"
    if dataset.startswith("Genotypic_data_(DArTAG"):
        return "dartag_panel2", "DArTAG Panel 2", "DArTAG"
    if dataset.startswith("Genotypic_data_from_CIMMYT"):
        return "cimmyt_bread_gbs_2013_2018", "GBS HapMap", "GBS"
    if dataset.startswith("Haplotype-based"):
        return "eyt_haplotype_blocks_2011_2018", "Haplotype blocks", "GBS-derived haplotypes"
    if dataset.startswith("IWYP64"):
        return "hibap35k", "35K breeders array", "SNP array"
    if dataset.startswith("Seeds_of_Discovery"):
        return "seeds_of_discovery_dartseq", "DArTseq", "DArTseq"
    return "unclassified", "unknown", "unknown"


def classify_file(relative_path: str) -> tuple[str, str]:
    rel = normalize_path(relative_path)
    lower = rel.lower()
    suffix = Path(lower).suffix
    if lower.endswith("desktop.ini"):
        return "system_metadata", "UNSUPPORTED_SYSTEM_FILE"
    if suffix == ".pdf" or any(token in lower for token in ("readme", "dictionary", "protocol", "agreement", "manifest.txt")):
        return "documentation", "DOCUMENTATION_ONLY"
    if lower.endswith("md5sum_hash.txt"):
        return "integrity_manifest", "PARSED_INTEGRITY_REFERENCE"
    if "doi" in lower:
        return "germplasm_doi_metadata", "PARSED_IDENTITY_EVIDENCE"
    if "sampleidvsgid" in lower or "samples_for_germinate" in lower:
        return "sample_gid_crosswalk", "PARSED_IDENTITY_EVIDENCE"
    if "germplasm" in lower or lower.endswith("sample.info.txt"):
        return "germplasm_metadata", "PARSED_SAMPLE_MEMBERSHIP"
    if "uchrom" in lower:
        return "marker_map", "PARSED_MARKER_MAP_PROFILE"
    if "phenotype" in lower:
        return "phenotype_support", "UNRELATED_TO_GENOTYPE_LINKAGE"
    if suffix in {".csv", ".txt", ".tab", ".flapjack"} or suffix == ".xlsx":
        if "fj_" in lower or "flapjack" in lower:
            return "alternate_marker_matrix", "ALTERNATE_REPRESENTATION_PROFILED"
        return "marker_matrix_or_results", "PARSED_SAMPLE_OR_MARKER_PROFILE"
    return "unclassified", "UNSUPPORTED_SOURCE"


class SampleLedger:
    def __init__(self) -> None:
        self.rows: dict[tuple[str, str], dict[str, object]] = {}
        self.evidence: list[dict[str, object]] = []

    def add(
        self,
        *,
        panel_id: str,
        raw_sample_id: object,
        source_file: str,
        source_location: str,
        identifier_type: str,
        candidate_gid: str = "",
        evidence_tier: str = "",
        evidence_type: str = "",
        normalization_rule: str = "PRESERVE_EXACT_STRING",
        marker_present: bool = False,
        metadata_present: bool = False,
        imputed_present: bool = False,
        kernel_order_present: bool = False,
        existing_qc: str = "NOT_ESTABLISHED",
        audit_qc: str = "NOT_ASSESSED_NO_FROZEN_THRESHOLD",
        matrix_order: int | None = None,
        metadata: dict[str, object] | None = None,
        accepted: bool = True,
        candidate_only: bool = False,
        sample_instance_key_value: str = "",
        mapping_status_override: str = "",
        join_rule: str = "",
        replicate_status: str = "",
        parser_version: str = VERSION,
    ) -> None:
        raw = clean(raw_sample_id)
        if not raw:
            return
        instance_key = clean(sample_instance_key_value) or panel_sample_key(panel_id, raw)
        key = (panel_id, instance_key)
        row = self.rows.setdefault(
            key,
            {
                "panel_id": panel_id,
                "panel_sample_key": instance_key,
                "sample_instance_key": instance_key,
                "raw_sample_id": raw,
                "raw_identifier_type": identifier_type,
                "candidate_gids": set(),
                "accepted_gid_evidence": set(),
                "source_files": set(),
                "source_locations": set(),
                "normalization_rules": set(),
                "evidence_tiers": set(),
                "evidence_types": set(),
                "marker_vector_present": False,
                "metadata_present": False,
                "existing_imputed_matrix_present": False,
                "existing_kernel_order_present": False,
                "existing_qc_statuses": set(),
                "audit_qc_statuses": set(),
                "matrix_orders": [],
                "metadata_json": [],
                "candidate_only": False,
                "mapping_status_overrides": set(),
                "join_rules": set(),
                "replicate_statuses": set(),
                "parser_versions": set(),
            },
        )
        row["source_files"].add(normalize_path(source_file))
        row["source_locations"].add(source_location)
        row["normalization_rules"].add(normalization_rule)
        row["marker_vector_present"] = bool(row["marker_vector_present"] or marker_present)
        row["metadata_present"] = bool(row["metadata_present"] or metadata_present)
        row["existing_imputed_matrix_present"] = bool(row["existing_imputed_matrix_present"] or imputed_present)
        row["existing_kernel_order_present"] = bool(row["existing_kernel_order_present"] or kernel_order_present)
        row["existing_qc_statuses"].add(existing_qc)
        row["audit_qc_statuses"].add(audit_qc)
        if mapping_status_override:
            row["mapping_status_overrides"].add(mapping_status_override)
        if join_rule:
            row["join_rules"].add(join_rule)
        if replicate_status:
            row["replicate_statuses"].add(replicate_status)
        row["parser_versions"].add(parser_version)
        if matrix_order is not None:
            row["matrix_orders"].append(int(matrix_order))
        if metadata:
            row["metadata_json"].append({str(k): clean(v) for k, v in metadata.items() if clean(v)})
        if candidate_gid:
            row["candidate_gids"].add(candidate_gid)
            if accepted and not candidate_only:
                row["accepted_gid_evidence"].add(candidate_gid)
            if candidate_only:
                row["candidate_only"] = True
            row["evidence_tiers"].add(evidence_tier)
            row["evidence_types"].add(evidence_type)
            self.evidence.append(
                {
                    "panel_id": panel_id,
                    "panel_sample_key": instance_key,
                    "sample_instance_key": instance_key,
                    "raw_sample_id": raw,
                    "candidate_canonical_gid": candidate_gid,
                    "evidence_tier": evidence_tier,
                    "evidence_type": evidence_type,
                    "accepted_by_rule": bool(accepted and not candidate_only),
                    "membership_by_rule": bool(accepted and not candidate_only),
                    "candidate_only": bool(candidate_only),
                    "source_file": normalize_path(source_file),
                    "source_location": source_location,
                    "normalization_rule": normalization_rule,
                }
            )

    def add_membership_evidence(
        self,
        *,
        panel_id: str,
        raw_identifier: str,
        candidate_gid: str,
        source_file: str,
        source_location: str,
        evidence_type: str,
        normalization_rule: str,
    ) -> None:
        if not candidate_gid:
            return
        evidence_key = f"{panel_id}::METADATA_MEMBERSHIP::{raw_identifier}"
        self.evidence.append(
            {
                "panel_id": panel_id,
                "panel_sample_key": evidence_key,
                "sample_instance_key": evidence_key,
                "raw_sample_id": raw_identifier,
                "candidate_canonical_gid": candidate_gid,
                "evidence_tier": "authoritative_explicit_gid_field",
                "evidence_type": evidence_type,
                "accepted_by_rule": False,
                "membership_by_rule": True,
                "candidate_only": False,
                "source_file": normalize_path(source_file),
                "source_location": source_location,
                "normalization_rule": normalization_rule,
            }
        )

    def finalize(self) -> pd.DataFrame:
        output: list[dict[str, object]] = []
        for key in sorted(self.rows):
            source = self.rows[key]
            accepted, resolution, ambiguity = resolve_gid_candidates(source["accepted_gid_evidence"])
            candidates = sorted(source["candidate_gids"])
            overrides = sorted(source["mapping_status_overrides"])
            if len(overrides) > 1:
                raise RuntimeError(f"Multiple mapping-status overrides for {source['panel_sample_key']}: {overrides}")
            if overrides:
                mapping = overrides[0]
            elif len(source["accepted_gid_evidence"]) > 1:
                mapping = "CONFLICTING_EVIDENCE"
                accepted = ""
            elif accepted:
                mapping = "ACCEPTED_EXPLICIT_TYPED_GID" if "authoritative_explicit_gid_field" in source["evidence_tiers"] else "ACCEPTED_AUTHORITATIVE_CROSSWALK"
            elif candidates:
                mapping = "CANDIDATE_REQUIRES_REVIEW" if len(candidates) == 1 else "AMBIGUOUS_MULTIPLE_GIDS"
            elif re.fullmatch(r"[0-9]+", str(source["raw_sample_id"])):
                mapping = "UNTYPED_NUMERIC_IDENTIFIER"
            else:
                mapping = "NO_CANONICAL_MATCH"
            marker = bool(source["marker_vector_present"])
            existing_qc_values = sorted(source["existing_qc_statuses"] - {"NOT_ESTABLISHED"})
            existing_qc = ";".join(existing_qc_values) if existing_qc_values else "NOT_ESTABLISHED"
            audit_qc = ";".join(sorted(source["audit_qc_statuses"]))
            if not marker:
                kernel = "MARKER_VECTOR_NOT_FOUND"
                reason = "sample is metadata-only or not found in an audited marker order"
            elif not accepted:
                kernel = "IDENTITY_NOT_VERIFIED"
                reason = "no accepted typed link to one canonical GID"
            elif "PASS" in existing_qc:
                kernel = "STRICT_KERNEL_READY_EXISTING_QC"
                reason = ""
            else:
                kernel = "MARKER_QC_NOT_ESTABLISHED"
                reason = "identity and marker presence verified but no frozen sample-QC contract"
            output.append(
                {
                    "panel_id": source["panel_id"],
                    "panel_sample_key": source["panel_sample_key"],
                    "sample_instance_key": source["sample_instance_key"],
                    "raw_sample_id": source["raw_sample_id"],
                    "raw_identifier_type": source["raw_identifier_type"],
                    "accepted_canonical_gid": accepted,
                    "mapping_status": mapping,
                    "evidence_tier": ";".join(sorted(source["evidence_tiers"])),
                    "evidence_type": ";".join(sorted(source["evidence_types"])),
                    "source_files": ";".join(sorted(source["source_files"])),
                    "source_locations": ";".join(sorted(source["source_locations"])),
                    "normalization_rules": ";".join(sorted(source["normalization_rules"])),
                    "ambiguity_set": ambiguity or ";".join(candidates),
                    "conflict_status": "CONFLICT" if mapping in {"CONFLICTING_EVIDENCE", "AMBIGUOUS_MULTIPLE_GIDS"} else "NONE",
                    "metadata_present": bool(source["metadata_present"]),
                    "marker_vector_present": marker,
                    "existing_imputed_matrix_present": bool(source["existing_imputed_matrix_present"]),
                    "existing_kernel_order_present": bool(source["existing_kernel_order_present"]),
                    "existing_qc_status": existing_qc,
                    "audit_qc_status": audit_qc,
                    "kernel_readiness_status": kernel,
                    "exclusion_or_unresolved_reason": reason,
                    "matrix_order_first": min(source["matrix_orders"]) if source["matrix_orders"] else pd.NA,
                    "matrix_occurrences": len(source["matrix_orders"]),
                    "raw_metadata_json": json.dumps(source["metadata_json"], ensure_ascii=False, sort_keys=True),
                    "join_rule": ";".join(sorted(source["join_rules"])),
                    "replicate_status": ";".join(sorted(source["replicate_statuses"])) or "NOT_ASSESSED",
                    "parser_version": ";".join(sorted(source["parser_versions"])),
                }
            )
        return pd.DataFrame(output)


def inventory_files(raw_inventory: Path, prior_profile: Path, genotype_root: Path) -> pd.DataFrame:
    inventory = pd.read_csv(raw_inventory, sep="\t", dtype=str, keep_default_na=False)
    profile = pd.read_csv(prior_profile, dtype=str, keep_default_na=False)
    profile["relative_norm"] = profile["relative_path"].map(normalize_path)
    line_counts = profile.set_index("relative_norm")["line_count_including_header"].to_dict()
    rows: list[dict[str, object]] = []
    for source in inventory.itertuples(index=False):
        rel = normalize_path(source.relative_path)
        path = genotype_root / Path(rel)
        panel_id, platform, technology = panel_id_for_file(rel)
        role, disposition = classify_file(rel)
        rows.append(
            {
                "source_root": source.source_root,
                "dataset": source.dataset,
                "relative_path": rel,
                "absolute_path": str(path.resolve()),
                "sha256": source.sha256,
                "bytes": int(source.bytes),
                "mtime_ns": int(source.mtime_ns),
                "format": Path(rel).suffix.lower().lstrip(".") or "none",
                "compression": "none",
                "panel_id": panel_id,
                "platform": platform,
                "technology": technology,
                "file_role": role,
                "prior_hash_bound_line_count": int(float(line_counts[rel])) if clean(line_counts.get(rel, "")) else pd.NA,
                "sample_orientation": "TO_BE_PROFILED" if "matrix" in role or "results" in role else "not_applicable",
                "raw_sample_count": pd.NA,
                "marker_count": pd.NA,
                "reference_genome_build": "NOT_DOCUMENTED_IN_FILE_INVENTORY",
                "allele_or_dosage_coding": "PROFILED_SEPARATELY",
                "missing_value_representation": "PROFILED_SEPARATELY",
                "ploidy_assumption": "NOT_INFERRED",
                "manifest_available": False,
                "existing_qc_documentation": "NOT_ESTABLISHED",
                "parse_status": disposition,
                "terminal_inventory_disposition": disposition,
            }
        )
    result = pd.DataFrame(rows)
    if len(result) != 92:
        raise RuntimeError(f"Expected 92 raw genotype files, observed {len(result)}")
    if result["relative_path"].duplicated().any():
        raise RuntimeError("Raw genotype inventory has duplicate paths")
    missing = [path for path in result["absolute_path"] if not Path(path).exists()]
    if missing:
        raise RuntimeError(f"Raw genotype files missing: {missing[:3]}")
    return result


def read_crosswalk(path: Path, sample_column: str, gid_column: str, delimiter: str = "\t") -> tuple[dict[str, set[str]], list[dict[str, str]]]:
    mapping: dict[str, set[str]] = defaultdict(set)
    evidence: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames or sample_column not in reader.fieldnames or gid_column not in reader.fieldnames:
            raise RuntimeError(f"Expected {sample_column}/{gid_column} in {path}: {reader.fieldnames}")
        for row_number, row in enumerate(reader, start=2):
            sample = clean(row.get(sample_column))
            gid, rule = authoritative_gid(row.get(gid_column), "explicit_crosswalk_gid_target")
            if sample and gid:
                mapping[sample].add(gid)
                evidence.append({"sample": sample, "gid": gid, "row": str(row_number), "rule": rule})
    return mapping, evidence


def add_hmp_reference(ledger: SampleLedger, order_path: Path) -> None:
    order = pd.read_csv(order_path, sep="\t", dtype=str, keep_default_na=False)
    sample_column = "sample_id" if "sample_id" in order else order.columns[0]
    for index, raw in enumerate(order[sample_column], start=1):
        gid, rule = authoritative_gid(raw, "frozen_gid_order")
        if not gid:
            raise RuntimeError(f"Non-GID value in frozen HMP order: {raw!r}")
        ledger.add(
            panel_id="frozen_hmp_v1",
            raw_sample_id=raw,
            source_file=str(order_path),
            source_location=f"row={index + 1};column={sample_column}",
            identifier_type="FROZEN_CANONICAL_GID_ORDER",
            candidate_gid=gid,
            evidence_tier="verified_frozen_kernel_sample_to_gid_provenance",
            evidence_type="frozen_hmp_kernel_order_gid",
            normalization_rule=rule,
            marker_present=True,
            metadata_present=True,
            kernel_order_present=True,
            existing_qc="PASS_FROZEN_HMP_QC",
            audit_qc="NOT_RECOMPUTED_IMMUTABLE_CERTIFIED_ORDER",
            matrix_order=index,
        )


def add_dartag(ledger: SampleLedger, root: Path) -> None:
    panel = "dartag_panel2"
    directory = root / "Genotypic_data_(DArTAG_panel_2)_for_the_IBWSN_and_SAWSN"
    matrix_specs = [
        (directory / "DArTAG_numeric.csv", False),
        (directory / "DArTAG_2moreOrders_numeric.csv", True),
    ]
    for path, subject_header in matrix_specs:
        rows = first_noncomment_rows(path, ",", 2 if subject_header else 1)
        if subject_header:
            samples = rows[0][1:]
            gids = rows[1][1:]
            if clean(rows[0][0]).upper() != "SUBJECT_ID" or clean(rows[1][0]).upper() != "GID":
                raise RuntimeError(f"Unexpected DArTAG two-row header: {path}")
        else:
            if clean(rows[0][0]).upper() != "GID":
                raise RuntimeError(f"Unexpected DArTAG GID header: {path}")
            gids = rows[0][1:]
            samples = [f"GID{clean(value)}" for value in gids]
        if len(samples) != len(gids):
            raise RuntimeError(f"DArTAG sample/GID cardinality mismatch: {path}")
        for index, (sample, raw_gid) in enumerate(zip(samples, gids), start=1):
            gid, rule = authoritative_gid(raw_gid, "documented_gid_row")
            ledger.add(
                panel_id=panel,
                raw_sample_id=sample,
                source_file=str(path),
                source_location=f"header_sample_column={index + 1};gid_row={'2' if subject_header else '1'}",
                identifier_type="PANEL_SUBJECT_ID" if subject_header else "DOCUMENTED_GID_AS_SAMPLE_LABEL",
                candidate_gid=gid,
                evidence_tier="authoritative_explicit_gid_field",
                evidence_type="documented_dartag_gid_row",
                normalization_rule=rule,
                marker_present=True,
                metadata_present=True,
                existing_qc="NOT_ESTABLISHED",
                matrix_order=index,
            )

    for filename in ("germplasm_list.xlsx", "germplasm_list_2.xlsx"):
        path = directory / filename
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header_row = 2 if filename == "germplasm_list.xlsx" else 1
        values = list(sheet.iter_rows(values_only=True))
        header = {clean(value).upper(): index for index, value in enumerate(values[header_row - 1]) if clean(value)}
        gid_column = header["GID"]
        subject_column = header.get("SUBJECT_ID")
        for row_number, row in enumerate(values[header_row:], start=header_row + 1):
            raw_gid = row[gid_column] if gid_column < len(row) else ""
            gid, rule = authoritative_gid(raw_gid, "authoritative_gid_column", excel=True)
            if not gid:
                continue
            raw_sample = row[subject_column] if subject_column is not None and subject_column < len(row) else gid
            metadata = {
                "CID": row[header["CID"]] if "CID" in header else "",
                "SID": row[header["SID"]] if "SID" in header else "",
                "pedigree": row[header.get("PEDIGREE", header.get("CROSS", -1))] if header.get("PEDIGREE", header.get("CROSS", -1)) >= 0 else "",
                "selection_history": row[header.get("SEL_HIST", -1)] if header.get("SEL_HIST", -1) >= 0 else "",
            }
            ledger.add(
                panel_id=panel,
                raw_sample_id=raw_sample,
                source_file=str(path),
                source_location=f"sheet={sheet.title};row={row_number};column=GID",
                identifier_type="PANEL_SUBJECT_ID" if subject_column is not None else "DOCUMENTED_GID_AS_SAMPLE_LABEL",
                candidate_gid=gid,
                evidence_tier="authoritative_explicit_gid_field",
                evidence_type="dartag_germplasm_manifest_gid",
                normalization_rule=rule,
                metadata_present=True,
                marker_present=False,
                metadata=metadata,
                existing_qc="NOT_ESTABLISHED",
            )
        workbook.close()


def add_hibap(ledger: SampleLedger, root: Path) -> pd.DataFrame:
    """Link physical HiBAP columns by matrix Entry number -> sidecar ENT.

    Matrix headers such as ``Hibap3`` and sidecar ``Sample 35k`` values are
    intentionally never joined.  Each physical marker column has its own stable
    sample-instance key, so duplicate entries and repeated GIDs remain separate.
    """

    instances, sidecar, summary = parse_hibap_sources(root)
    if summary != {
        **summary,
        "matrix_columns": 148,
        "matrix_header_to_sidecar_sample35k_agreement": 0,
        "entry_to_ent_agreement": 148,
        "unique_matrix_headers": 148,
        "unique_entry_numbers": 147,
        "unique_linked_gids": 145,
        "matrix_sidecar_gid_concordant": 148,
        "accepted_columns": 148,
        "gid_conflicts": 0,
        "duplicate_entry_109_columns": 2,
    }:
        raise RuntimeError(f"HiBAP source semantics differ from the frozen R2 contract: {summary}")
    for row in instances.to_dict("records"):
        accepted = bool(row["accepted_canonical_gid"])
        ledger.add(
            panel_id="hibap35k",
            raw_sample_id=row["raw_matrix_header"],
            source_file=row["source_file"],
            source_location=(
                f"header_row=4;entry_row=2;gid_row=3;sample_column={row['physical_column_index']};"
                f"sidecar_row={row['sidecar_physical_row']};sidecar_columns=ENT,GID,Sample 35k"
            ),
            identifier_type="HIBAP35K_MATRIX_SAMPLE_HEADER",
            candidate_gid=row["accepted_canonical_gid"] or row["matrix_canonical_gid"],
            evidence_tier="authoritative_explicit_gid_field" if accepted else "",
            evidence_type="hibap_entry_ent_and_gid_concordance" if accepted else "hibap_entry_ent_gid_conflict",
            normalization_rule="EXACT_ENTRY_ENT_JOIN_AND_TYPED_GID_CONCORDANCE",
            marker_present=True,
            metadata_present=True,
            existing_qc="NOT_ESTABLISHED",
            matrix_order=int(row["matrix_order"]),
            metadata={
                "HIBAP35K_MATRIX_SAMPLE_HEADER": row["raw_matrix_header"],
                "HIBAP35K_MATRIX_ENTRY_NUMBER": row["matrix_entry_number"],
                "HIBAP35K_MATRIX_GID": row["matrix_gid_raw"],
                "HIBAP35K_SIDECAR_ENT": row["sidecar_ent"],
                "HIBAP35K_SIDECAR_GID": row["sidecar_gid_raw"],
                "HIBAP35K_SIDECAR_SAMPLE_35K": row["sidecar_sample_35k"],
                "CID": row["sidecar_cid"],
                "SID": row["sidecar_sid"],
                "pedigree": row["sidecar_cross_name"],
                "selection_history": row["sidecar_selection_history"],
            },
            accepted=accepted,
            candidate_only=False,
            sample_instance_key_value=row["sample_instance_key"],
            mapping_status_override=row["linkage_status"],
            join_rule=row["join_rule"],
            replicate_status=row["replicate_status"],
            parser_version=R2_PARSER_VERSION,
        )

    # Preserve the complete typed sidecar GID set as panel-membership evidence.
    # It is not a sample crosswalk and cannot create a marker-linked identity.
    for row in sidecar.to_dict("records"):
        ledger.add_membership_evidence(
            panel_id="hibap35k",
            raw_identifier=f"SIDECAR_ENT={row['ENT']}",
            candidate_gid=row["canonical_gid"],
            source_file=row["sidecar_source_file"],
            source_location=f"row={row['sidecar_physical_row']};columns=ENT,GID",
            evidence_type="hibap_sidecar_typed_gid_metadata_membership_only",
            normalization_rule="TYPED_SIDECAR_GID_RETAINED_SEPARATE_FROM_SAMPLE_LINKAGE",
        )
    return instances


def add_simple_header_gbs(ledger: SampleLedger, root: Path) -> None:
    for matrix_path in sorted((root / "GBS").rglob("*.txt")):
        if matrix_path.name.upper() == "MANIFEST.TXT":
            continue
        panel, _, _ = panel_id_for_file(str(matrix_path.relative_to(root)))
        header = next(iter(delimited_rows(matrix_path, "\t", limit=1)))
        if [clean(value).upper() for value in header[:4]] != ["S", "PRESENT", "MAF", "PERCENTHET"]:
            continue
        for index, raw in enumerate(header[4:], start=1):
            gid, rule = authoritative_gid(raw, "documented_gid_row")
            ledger.add(
                panel_id=panel,
                raw_sample_id=raw,
                source_file=str(matrix_path),
                source_location=f"header_row=1;sample_column={index + 4}",
                identifier_type="DOCUMENTED_GID_MATRIX_HEADER",
                candidate_gid=gid,
                evidence_tier="authoritative_explicit_gid_field",
                evidence_type="gbs_matrix_gid_header",
                normalization_rule=rule,
                marker_present=True,
                metadata_present=True,
                existing_qc="NOT_ESTABLISHED",
                matrix_order=index,
            )


def add_sample_crosswalk_panel(
    ledger: SampleLedger,
    *,
    panel_id: str,
    crosswalk_path: Path,
    sample_column: str,
    gid_column: str,
    matrix_path: Path,
) -> dict[str, set[str]]:
    mapping, evidence = read_crosswalk(crosswalk_path, sample_column, gid_column)
    for item in evidence:
        ledger.add(
            panel_id=panel_id,
            raw_sample_id=item["sample"],
            source_file=str(crosswalk_path),
            source_location=f"row={item['row']};columns={sample_column},{gid_column}",
            identifier_type="PANEL_SAMPLE_ID",
            candidate_gid=item["gid"],
            evidence_tier="authoritative_explicit_gid_field",
            evidence_type="explicit_sample_to_gid_sidecar",
            normalization_rule=item["rule"],
            metadata_present=True,
            existing_qc="NOT_ESTABLISHED",
        )
    header = next(iter(delimited_rows(matrix_path, "\t", limit=1)))
    if clean(header[0]).upper() != "MARKERID":
        raise RuntimeError(f"Unexpected marker-by-sample header: {matrix_path}")
    for index, sample in enumerate(header[1:], start=1):
        candidates = mapping.get(clean(sample), set())
        for gid in sorted(candidates) or [""]:
            ledger.add(
                panel_id=panel_id,
                raw_sample_id=sample,
                source_file=str(matrix_path),
                source_location=f"header_row=1;sample_column={index + 1}",
                identifier_type="PANEL_SAMPLE_ID",
                candidate_gid=gid,
                evidence_tier="authoritative_explicit_gid_field" if gid else "",
                evidence_type="exact_sample_id_to_same_panel_sidecar" if gid else "",
                normalization_rule="EXACT_STRING_CROSSWALK_LOOKUP" if gid else "PRESERVE_EXACT_OPAQUE_STRING",
                marker_present=True,
                metadata_present=bool(gid),
                existing_qc="NOT_ESTABLISHED",
                matrix_order=index,
            )
    return mapping


def add_mexican_results_gid_rows(ledger: SampleLedger, root: Path) -> None:
    directory = root / "DArTseq-derived_SNPs_for_wheat_Mexican_landrace_accessions"
    for path in sorted(directory.glob("Mexican_Land_Races_results_*.csv")):
        rows = list(delimited_rows(path, ",", limit=7))
        sample_row = rows[0]
        gid_row = rows[6]
        try:
            sample_start = next(index for index, value in enumerate(sample_row) if clean(value).upper() == "SAMPLEID") + 1
            gid_start = next(index for index, value in enumerate(gid_row) if clean(value).upper() == "GID") + 1
        except StopIteration as exc:
            raise RuntimeError(f"Missing SampleID/GID preamble rows in {path}") from exc
        if sample_start != gid_start:
            raise RuntimeError(f"Mexican results sample/GID columns do not align in {path}")
        samples = sample_row[sample_start:]
        gids = gid_row[gid_start:]
        if len(samples) != len(gids):
            raise RuntimeError(f"Mexican results sample/GID count mismatch in {path}")
        for index, (sample, raw_gid) in enumerate(zip(samples, gids), start=1):
            gid, rule = authoritative_gid(raw_gid, "documented_gid_row")
            if not gid:
                continue
            ledger.add(
                panel_id="mexican_landrace_dartseq",
                raw_sample_id=sample,
                source_file=str(path),
                source_location=f"sample_row=1;gid_row=7;sample_column={sample_start + index}",
                identifier_type="PANEL_SAMPLE_ID_WITH_PARALLEL_GID_ROW",
                candidate_gid=gid,
                evidence_tier="authoritative_explicit_gid_field",
                evidence_type="mexican_results_parallel_gid_row",
                normalization_rule=rule,
                marker_present=True,
                metadata_present=True,
                existing_qc="NOT_ESTABLISHED",
                matrix_order=index,
            )


def add_80k_panels(ledger: SampleLedger, root: Path, candidate_maps: list[tuple[str, dict[str, set[str]]]]) -> pd.DataFrame:
    """Use one canonical PAV sample axis per population without collapsing labels."""

    axes = canonical_80k_pav_axes(root)
    for panel, panel_axis in axes.groupby("panel_id", sort=True):
        panel_axis = panel_axis.sort_values("physical_column_index")
        for matrix_order, row in enumerate(panel_axis.to_dict("records"), start=1):
            sample = row["raw_sample_label"]
            candidates: set[str] = set()
            sources: list[str] = []
            for source_name, mapping in candidate_maps:
                if sample in mapping:
                    candidates.update(mapping[sample])
                    sources.append(source_name)
            ledger.add(
                panel_id=panel,
                raw_sample_id=sample,
                source_file=row["source_file"],
                source_location=(
                    f"sample_id_preamble_row={row['sample_id_physical_row']};"
                    f"sample_column={row['physical_column_index']};occurrence={row['occurrence_index']}"
                ),
                identifier_type="OPAQUE_PANEL_SAMPLE_ID",
                marker_present=True,
                metadata_present=False,
                existing_qc="NOT_ESTABLISHED",
                matrix_order=matrix_order,
                normalization_rule="PRESERVE_EXACT_OPAQUE_STRING",
                metadata={
                    "well": row["well"],
                    "plate_or_barcode": row["plate_or_barcode"],
                    "sample_group": row["sample_group"],
                    "replicate_or_index": row["replicate_or_index"],
                    "physical_column_index": row["physical_column_index"],
                    "occurrence_index": row["occurrence_index"],
                },
                sample_instance_key_value=row["sample_instance_key"],
                join_rule="PRESERVE_CANONICAL_PAV_PHYSICAL_SAMPLE_AXIS_NO_IDENTITY_INFERENCE",
                replicate_status="DUPLICATE_LABEL_OCCURRENCE_RETAINED" if int(row["occurrence_index"]) > 1 else "UNIQUE_LABEL_OCCURRENCE",
                parser_version=R2_PARSER_VERSION,
            )
            for gid in sorted(candidates):
                ledger.add(
                    panel_id=panel,
                    raw_sample_id=sample,
                    source_file=";".join(sources),
                    source_location="exact_identifier_in_different_panel_crosswalk",
                    identifier_type="OPAQUE_PANEL_SAMPLE_ID",
                    candidate_gid=gid,
                    evidence_tier="candidate_cross_panel_exact_sample_label",
                    evidence_type="cross_panel_exact_sample_id_candidate_only",
                    normalization_rule="EXACT_STRING_ONLY_NO_NAMESPACE_COLLAPSE",
                    marker_present=True,
                    accepted=False,
                    candidate_only=True,
                    existing_qc="NOT_ESTABLISHED",
                    sample_instance_key_value=row["sample_instance_key"],
                    join_rule="CROSS_PANEL_EXACT_LABEL_CANDIDATE_ONLY",
                    replicate_status="DUPLICATE_LABEL_OCCURRENCE_RETAINED" if int(row["occurrence_index"]) > 1 else "UNIQUE_LABEL_OCCURRENCE",
                    parser_version=R2_PARSER_VERSION,
                )
    return axes


def add_cimmyt_bread_panel(ledger: SampleLedger, root: Path) -> None:
    directory = root / "Genotypic_data_from_CIMMYT_bread_wheat_breeding_lines"
    marker_path = next(directory.glob("*.hmp.txt"))
    header = next(iter(delimited_rows(marker_path, "\t", limit=1)))
    for index, raw in enumerate(header[11:], start=1):
        gid, rule = authoritative_gid(raw, "documented_gid_row")
        ledger.add(
            panel_id="cimmyt_bread_gbs_2013_2018",
            raw_sample_id=raw,
            source_file=str(marker_path),
            source_location=f"header_row=1;sample_column={index + 11}",
            identifier_type="DOCUMENTED_GERMPLASM_ID_MATRIX_HEADER",
            candidate_gid=gid,
            evidence_tier="authoritative_explicit_gid_field" if gid else "",
            evidence_type="documented_hapmap_germplasm_id" if gid else "",
            normalization_rule=rule if gid else "PRESERVE_INVALID_OR_OPAQUE_SAMPLE_NAME",
            marker_present=True,
            metadata_present=bool(gid),
            imputed_present=True,
            existing_qc="PASS_EXISTING_MAF0.01_MISS50_HET10_IMPUTED_EXPORT",
            audit_qc="NOT_RECOMPUTED_EXISTING_EXPORT_ONLY",
            matrix_order=index,
        )

    sample_info = directory / "sample.info.txt"
    rows = delimited_rows(sample_info, "\t")
    header_info = next(rows)
    if [clean(value).upper() for value in header_info[:2]] != ["TA", "SUBSPECIES"]:
        raise RuntimeError("Unexpected sample.info.txt schema")
    for row_number, row in enumerate(rows, start=2):
        raw = row[0] if row else ""
        if not clean(raw):
            continue
        # TA is explicitly a sample/taxa namespace.  A GID-looking prefix or
        # numeric suffix is not accepted without a TA-to-GID crosswalk.
        ledger.add(
            panel_id="cimmyt_bread_gbs_2013_2018_ta_metadata",
            raw_sample_id=raw,
            source_file=str(sample_info),
            source_location=f"row={row_number};column=TA",
            identifier_type="OPAQUE_TAXA_ID",
            metadata_present=True,
            marker_present=False,
            normalization_rule="PRESERVE_EXACT_OPAQUE_STRING",
            metadata={"subspecies": row[1] if len(row) > 1 else ""},
            existing_qc="NOT_ESTABLISHED",
        )


def add_haplotype_panel(ledger: SampleLedger, root: Path) -> None:
    path = root / "Haplotype-based_genome-wide_association_study" / "Haplotype_blocks_EYT2011-12_to_EYT2017-18.csv"
    rows = delimited_rows(path, ",")
    header = next(rows)
    columns = {clean(value).upper(): index for index, value in enumerate(header)}
    if not {"GID", "EYT"}.issubset(columns):
        raise RuntimeError("Haplotype block table lacks GID/EYT")
    for row_number, row in enumerate(rows, start=2):
        raw_gid = row[columns["GID"]] if columns["GID"] < len(row) else ""
        eyt = row[columns["EYT"]] if columns["EYT"] < len(row) else ""
        gid, rule = authoritative_gid(raw_gid, "authoritative_gid_column")
        if not gid:
            continue
        raw_sample = f"{gid}|EYT={clean(eyt)}"
        ledger.add(
            panel_id="eyt_haplotype_blocks_2011_2018",
            raw_sample_id=raw_sample,
            source_file=str(path),
            source_location=f"row={row_number};columns=GID,EYT",
            identifier_type="GID_BY_EYT_PANEL_ROW",
            candidate_gid=gid,
            evidence_tier="authoritative_explicit_gid_field",
            evidence_type="haplotype_table_gid_column",
            normalization_rule=rule,
            marker_present=True,
            metadata_present=True,
            existing_qc="NOT_ESTABLISHED",
            matrix_order=row_number - 1,
            metadata={"EYT": eyt},
        )


def add_mas_xlsx_panels(ledger: SampleLedger, root: Path) -> None:
    specs = [
        (
            "mas_57ibwsn_42sawsn_35hrwsn",
            root / "57IBWSN,_42SAWSN,_and_35HRWSN_-_Gene-based_marker_data_for_marker-assisted_selection" / "57IBWSN,_42SAWSN,_35HRWSN_results.xlsx",
            "Results",
        ),
        (
            "mas_58ibwsn_43sawsn",
            root / "58IBWSN_and_43SAWSN_-_Gene-based_marker_data_for_marker-assisted_selection" / "58IBWSN-43SAWSN_results.xlsx",
            "BW24GSSD-B01 Sample Info",
        ),
    ]
    for panel, path, sheet_name in specs:
        header_row, columns = xlsx_header(path, sheet_name, {"SAMPLEID", "GID"})
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            sample = row[columns["SAMPLEID"]] if columns["SAMPLEID"] < len(row) else ""
            raw_gid = row[columns["GID"]] if columns["GID"] < len(row) else ""
            gid, rule = authoritative_gid(raw_gid, "authoritative_gid_column", excel=True)
            ledger.add(
                panel_id=panel,
                raw_sample_id=sample,
                source_file=str(path),
                source_location=f"sheet={sheet_name};row={row_number};columns=SampleID,GID",
                identifier_type="PANEL_SAMPLE_ID",
                candidate_gid=gid,
                evidence_tier="authoritative_explicit_gid_field" if gid else "",
                evidence_type="mas_results_gid_column" if gid else "",
                normalization_rule=rule if gid else "PRESERVE_EXACT_OPAQUE_STRING",
                marker_present=True,
                metadata_present=True,
                existing_qc="NOT_ESTABLISHED",
                matrix_order=row_number - header_row,
                metadata={
                    "pedigree": row[columns["PEDIGREE"]] if "PEDIGREE" in columns and columns["PEDIGREE"] < len(row) else "",
                    "nursery": row[columns["NURSERY"]] if "NURSERY" in columns and columns["NURSERY"] < len(row) else "",
                },
            )
        workbook.close()


def add_45ibwsn_mas(ledger: SampleLedger, root: Path) -> None:
    path = root / "GBS" / "45th_International_Bread_Wheat_Screening_Nursery_MAS_data" / "45IBWSN.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Field Book"]
    values = list(sheet.iter_rows(values_only=True))
    header_index = next(index for index, row in enumerate(values) if {clean(value).upper() for value in row} >= {"CID", "SID", "GID"})
    columns = {clean(value).upper(): index for index, value in enumerate(values[header_index]) if clean(value)}
    for row_number, row in enumerate(values[header_index + 1 :], start=header_index + 2):
        gid, rule = authoritative_gid(row[columns["GID"]] if columns["GID"] < len(row) else "", "authoritative_gid_column", excel=True)
        cid = clean(row[columns["CID"]]) if columns["CID"] < len(row) else ""
        sid = clean(row[columns["SID"]]) if columns["SID"] < len(row) else ""
        entry = clean(row[columns["ENTRY"]]) if "ENTRY" in columns and columns["ENTRY"] < len(row) else ""
        raw_sample = f"CID={cid}|SID={sid}|ENTRY={entry}"
        ledger.add(
            panel_id="mas_45ibwsn",
            raw_sample_id=raw_sample,
            source_file=str(path),
            source_location=f"sheet=Field Book;row={row_number};column=GID",
            identifier_type="PANEL_ENTRY_COMPOSITE_KEY",
            candidate_gid=gid,
            evidence_tier="authoritative_explicit_gid_field" if gid else "",
            evidence_type="fieldbook_gid_column" if gid else "",
            normalization_rule=rule if gid else "PRESERVE_COMPOSITE_ENTRY_KEY",
            marker_present=True,
            metadata_present=True,
            existing_qc="NOT_ESTABLISHED",
            matrix_order=row_number - header_index - 1,
            metadata={
                "CID": cid,
                "SID": sid,
                "pedigree": row[columns["CROSS NAME"]] if "CROSS NAME" in columns and columns["CROSS NAME"] < len(row) else "",
                "selection_history": row[columns["SELECTION HISTORY"]] if "SELECTION HISTORY" in columns and columns["SELECTION HISTORY"] < len(row) else "",
            },
        )
    workbook.close()


def profile_genotype_files(inventory: pd.DataFrame, samples: pd.DataFrame) -> pd.DataFrame:
    source_sample_counts: dict[str, int] = defaultdict(int)
    for row in samples.itertuples(index=False):
        for value in str(row.source_files).split(";"):
            if value:
                # Source paths were already emitted as absolute paths. Avoid a
                # filesystem stat/resolve for every sample (hundreds of
                # thousands of redundant I/O operations on /mnt/e).
                source_sample_counts[normalize_path(value)] += 1

    output = inventory.copy()
    for index, row in output.iterrows():
        path = Path(row["absolute_path"])
        normalized_absolute = normalize_path(path)
        role = row["file_role"]
        lower = row["relative_path"].lower()
        sample_count = source_sample_counts.get(normalized_absolute, 0)
        marker_count: int | pd._libs.missing.NAType = pd.NA
        orientation = row["sample_orientation"]
        coding = row["allele_or_dosage_coding"]
        missing = row["missing_value_representation"]
        dimension_rule = "not_applicable"
        line_count = row["prior_hash_bound_line_count"]
        if role in {"marker_matrix_or_results", "alternate_marker_matrix"} and path.suffix.lower() in {".txt", ".csv", ".tab", ".flapjack"}:
            delimiter = "\t" if path.suffix.lower() in {".txt", ".tab", ".flapjack"} else ","
            if "intertek_format" in lower:
                header = next(iter(delimited_rows(path, delimiter, limit=1)))
                orientation = "sample_by_marker"
                marker_count = max(0, len(header) - 2)
                sample_count = max(sample_count, int(line_count) - 1 if pd.notna(line_count) else 0)
                coding = "platform allele calls"
                missing = "-:- or platform-specific blank"
                dimension_rule = "header_width_minus_two;hash_bound_line_count_minus_header"
            elif "dartag" in lower and "numeric" in lower:
                header_rows = first_noncomment_rows(path, ",", 2)
                has_subject = clean(header_rows[0][0]).upper() == "SUBJECT_ID"
                orientation = "marker_by_sample"
                sample_count = len(header_rows[0]) - 1
                marker_count = int(line_count) - (2 if has_subject else 1) if pd.notna(line_count) else pd.NA
                coding = "0/1/2 dosage"
                missing = "-"
                dimension_rule = "header_width_minus_one;hash_bound_line_count_minus_preamble"
            elif "snps_35karray" in lower:
                rows = first_noncomment_rows(path, "\t", 4)
                orientation = "marker_by_sample"
                sample_count = len(rows[3]) - 11
                marker_count = int(line_count) - 4 if pd.notna(line_count) else pd.NA
                coding = "IUPAC nucleotide calls"
                missing = "N"
                dimension_rule = "header_width_minus_11;hash_bound_line_count_minus_four"
            elif lower.endswith(".hmp.txt"):
                header = next(iter(delimited_rows(path, "\t", limit=1)))
                orientation = "marker_by_sample"
                sample_count = len(header) - 11
                marker_count = int(line_count) - 1 if pd.notna(line_count) else pd.NA
                coding = "IUPAC nucleotide calls"
                missing = "N, . or -"
                dimension_rule = "HapMap header width minus 11;hash-bound line count minus header"
            elif "sawytgbs" in lower or "sawyt_gbs.txt" in lower:
                header = next(iter(delimited_rows(path, "\t", limit=1)))
                orientation = "marker_by_sample"
                sample_count = len(header) - 4
                marker_count = int(line_count) - 1 if pd.notna(line_count) else pd.NA
                coding = "IUPAC nucleotide calls"
                missing = "N"
                dimension_rule = "header width minus four;hash-bound line count minus header"
            elif "haplotype_blocks" in lower:
                header = next(iter(delimited_rows(path, ",", limit=1)))
                orientation = "sample_by_marker"
                sample_count = int(line_count) - 1 if pd.notna(line_count) else sample_count
                marker_count = len(header) - 2
                coding = "haplotype allele strings"
                missing = "NA or N-containing code"
                dimension_rule = "hash-bound line count minus header;header width minus GID/EYT"
            elif any(token in lower for token in ("_fj_", "flapjack", "inverted_fj", "fj_format")):
                rows = first_noncomment_rows(path, "\t", 1)
                header = rows[0]
                orientation = "sample_by_marker"
                marker_count = len(header) - 1
                comment_rows = 2 if lower.startswith("80k/") else 0
                sample_count = int(line_count) - comment_rows - 1 if pd.notna(line_count) else sample_count
                coding = "nucleotide or 0/1 presence calls"
                missing = "-"
                dimension_rule = "first non-comment header width minus one;hash-bound line count minus preamble"
            elif "seq_snps_extract" in lower and path.suffix.lower() == ".txt":
                header = next(iter(delimited_rows(path, "\t", limit=1)))
                orientation = "marker_by_sample"
                sample_count = len(header) - 1
                marker_count = int(line_count) - 1 if pd.notna(line_count) else pd.NA
                coding = "nucleotide/IUPAC calls"
                missing = "-"
                dimension_rule = "header width minus one;hash-bound line count minus header"
            elif lower.startswith("80k/") and path.suffix.lower() == ".csv":
                rows = first_noncomment_rows(path, ",", 6)
                orientation = "marker_by_sample"
                sample_count = sum(clean(value) not in {"", "*"} for value in rows[4])
                marker_count = int(line_count) - 8 if pd.notna(line_count) else pd.NA
                coding = "nucleotide or 0/1 presence calls"
                missing = "-"
                dimension_rule = "typed sample preamble row;hash-bound line count minus two notices and six preamble rows"
            elif "seq_snps_extract" in lower and path.suffix.lower() == ".csv":
                rows = first_noncomment_rows(path, ",", 5)
                orientation = "marker_by_sample"
                metadata_width = next(index for index, value in enumerate(rows[-1]) if clean(value).upper() in {"SEEDSynt979".upper(), "SEEDMIX1552".upper()})
                sample_count = len(rows[-1]) - metadata_width
                marker_count = int(line_count) - 5 if pd.notna(line_count) else pd.NA
                dimension_rule = "schema header sample suffix;hash-bound line count minus preamble"
            elif "results_" in lower:
                orientation = "mixed_results_export"
                marker_count = int(line_count) - 8 if pd.notna(line_count) else pd.NA
                dimension_rule = "hash-bound result-row count;sample/GID parallel preamble"
        elif path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            if role == "phenotype_support":
                orientation = "phenotype_support_not_genotype_matrix"
            else:
                max_rows = max(workbook[name].max_row for name in workbook.sheetnames)
                max_columns = max(workbook[name].max_column for name in workbook.sheetnames)
                sample_count = max(sample_count, max_rows - 1)
                marker_count = max_columns
                orientation = "workbook_mixed_metadata_and_marker_results"
                dimension_rule = "workbook physical max rows/columns;semantic sample count from ledger"
            workbook.close()
        if sample_count:
            output.at[index, "raw_sample_count"] = int(sample_count)
        if pd.notna(marker_count):
            output.at[index, "marker_count"] = int(marker_count)
        output.at[index, "sample_orientation"] = orientation
        output.at[index, "allele_or_dosage_coding"] = coding
        output.at[index, "missing_value_representation"] = missing
        output.at[index, "dimension_derivation"] = dimension_rule
        panel_files = output[output["panel_id"].eq(row["panel_id"])]
        output.at[index, "manifest_available"] = bool(panel_files["file_role"].isin({"germplasm_metadata", "germplasm_doi_metadata", "sample_gid_crosswalk"}).any())
        if row["panel_id"] == "cimmyt_bread_gbs_2013_2018" and path.name.endswith("hmp.txt"):
            output.at[index, "existing_qc_documentation"] = "MAF0.01_Miss50_Het10;imputed;filename_and_readme"
        output.at[index, "parse_status"] = "PASS_TERMINAL_FILE_ACCOUNTING"
    return output


def build_namespace_collisions(samples: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for raw, group in samples.groupby("raw_sample_id", sort=True):
        panels = sorted(set(group["panel_id"]))
        if len(panels) > 1:
            rows.append(
                {
                    "collision_type": "SAME_RAW_SAMPLE_LABEL_ACROSS_PANELS",
                    "raw_identifier": raw,
                    "panels": ";".join(panels),
                    "panel_sample_keys": ";".join(sorted(group["panel_sample_key"])),
                    "accepted_gids": ";".join(sorted(set(group["accepted_canonical_gid"]) - {""})),
                    "disposition": "NAMESPACES_RETAINED_NO_CROSS_PANEL_IDENTITY_INFERENCE",
                }
            )
    opaque = samples[
        samples["raw_identifier_type"].str.contains("OPAQUE|TAXA", case=False, na=False)
        & samples["raw_sample_id"].str.fullmatch(r"(?:GID)?[0-9]+(?:_[0-9]+)?", case=False, na=False)
    ]
    for row in opaque.itertuples(index=False):
        rows.append(
            {
                "collision_type": "GID_LOOKING_OR_NUMERIC_OPAQUE_LABEL",
                "raw_identifier": row.raw_sample_id,
                "panels": row.panel_id,
                "panel_sample_keys": row.panel_sample_key,
                "accepted_gids": row.accepted_canonical_gid,
                "disposition": "NO_GID_FROM_OPAQUE_LABEL_PAYLOAD",
            }
        )
    return pd.DataFrame(rows).drop_duplicates() if rows else pd.DataFrame(columns=["collision_type", "raw_identifier", "panels", "panel_sample_keys", "accepted_gids", "disposition"])


def build_duplicate_report(samples: pd.DataFrame) -> pd.DataFrame:
    accepted = samples[samples["accepted_canonical_gid"].ne("")]
    rows: list[dict[str, object]] = []
    for gid, group in accepted.groupby("accepted_canonical_gid", sort=True):
        if len(group) < 2:
            continue
        panels = sorted(set(group["panel_id"]))
        rows.append(
            {
                "canonical_gid": gid,
                "panel_sample_count": len(group),
                "panel_count": len(panels),
                "panels": ";".join(panels),
                "panel_sample_keys": ";".join(sorted(group["panel_sample_key"])),
                "relationship_class": "CROSS_PANEL_REPRESENTATIONS" if len(panels) > 1 else "WITHIN_PANEL_POSSIBLE_TECHNICAL_REPLICATES",
                "marker_concordance_status": "NOT_COMPUTED_NO_FROZEN_HARMONIZED_COMMON_MARKER_SPACE",
                "identity_interpretation": "typed metadata establishes shared GID; samples remain distinct pending concordance",
                "recommended_policy": "retain panel-specific representations; assess marker concordance before merging",
            }
        )
    return pd.DataFrame(rows)


def write_orders(samples: pd.DataFrame, out_dir: Path) -> pd.DataFrame:
    order_dir = out_dir / "orders"
    order_dir.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, object]] = []
    for panel_id, group in samples.groupby("panel_id", sort=True):
        group = group.copy()
        group["_order"] = pd.to_numeric(group["matrix_order_first"], errors="coerce").fillna(10**12)
        group = group.sort_values(["_order", "raw_sample_id", "panel_sample_key"]).drop(columns="_order")
        sets = {
            "all_discovered": group,
            "identity_verified_marker_present": group[group["accepted_canonical_gid"].ne("") & group["marker_vector_present"]],
            "gid_replicate_resolution_required": group[
                group["accepted_canonical_gid"].ne("")
                & group["accepted_canonical_gid"].duplicated(keep=False)
            ],
        }
        strict = group[group["kernel_readiness_status"].eq("STRICT_KERNEL_READY_EXISTING_QC")]
        if not strict.empty:
            sets["strict_kernel_ready"] = strict
        for order_type, frame in sets.items():
            path = order_dir / f"{panel_id}__{order_type}.tsv"
            columns = [
                "panel_id", "panel_sample_key", "raw_sample_id", "accepted_canonical_gid",
                "matrix_order_first", "mapping_status", "marker_vector_present",
                "existing_qc_status", "kernel_readiness_status",
            ]
            frame[columns].to_csv(path, sep="\t", index=False)
            manifest.append(
                {
                    "panel_id": panel_id,
                    "order_type": order_type,
                    "rows": len(frame),
                    "relative_path": normalize_path(path.relative_to(out_dir)),
                    "sha256": sha256(path),
                }
            )
    return pd.DataFrame(manifest)


def stage1_linkage(
    stage1_path: Path,
    samples: pd.DataFrame,
    evidence: pd.DataFrame,
    out_dir: Path,
) -> dict[str, object]:
    con = duckdb.connect(database=":memory:")
    stage1_sql_path = str(stage1_path).replace("'", "''")
    con.execute(f"CREATE VIEW stage1 AS SELECT * FROM read_parquet('{stage1_sql_path}')")
    stage1_gids = set(
        "GID" + clean(value)
        for value in con.execute("SELECT DISTINCT resolved_gid FROM stage1").fetch_df()["resolved_gid"]
    )
    panel_states: list[dict[str, object]] = []
    coverage_rows: list[dict[str, object]] = []
    panel_gid_sets: dict[str, set[str]] = {}
    membership_gid_sets: dict[str, set[str]] = {}
    marker_gid_sets: dict[str, set[str]] = {}
    qc_gid_sets: dict[str, set[str]] = {}
    strict_gid_sets: dict[str, set[str]] = {}
    for panel_id, group in samples.groupby("panel_id", sort=True):
        accepted = set(group.loc[group["accepted_canonical_gid"].ne(""), "accepted_canonical_gid"])
        panel_evidence = evidence[
            evidence["panel_id"].eq(panel_id)
            & evidence["membership_by_rule"].fillna(False)
            & ~evidence["candidate_only"].fillna(False)
        ]
        membership = set(panel_evidence["candidate_canonical_gid"]) - {""}
        marker = set(group.loc[group["accepted_canonical_gid"].ne("") & group["marker_vector_present"], "accepted_canonical_gid"])
        qc = set(group.loc[group["accepted_canonical_gid"].ne("") & group["existing_qc_status"].str.contains("PASS", na=False), "accepted_canonical_gid"])
        strict = set(group.loc[group["kernel_readiness_status"].eq("STRICT_KERNEL_READY_EXISTING_QC"), "accepted_canonical_gid"])
        panel_gid_sets[panel_id] = accepted
        membership_gid_sets[panel_id] = membership | accepted
        marker_gid_sets[panel_id] = marker
        qc_gid_sets[panel_id] = qc
        strict_gid_sets[panel_id] = strict
        for gid in sorted(stage1_gids):
            coverage_rows.append(
                {
                    "canonical_gid": gid,
                    "panel_id": panel_id,
                    "metadata_or_membership": gid in membership_gid_sets[panel_id],
                    "accepted_gid_linkage": gid in accepted,
                    "raw_marker_vector_present": gid in marker,
                    "existing_qc_pass": gid in qc,
                    "strict_kernel_ready": gid in strict,
                }
            )
    coverage = pd.DataFrame(coverage_rows)
    coverage.to_csv(out_dir / "canonical_gid_panel_coverage.tsv", sep="\t", index=False)
    coverage.to_parquet(out_dir / "canonical_gid_panel_coverage.parquet", index=False)

    selected_trait_frame = pd.DataFrame({"accepted_canonical_trait": sorted(SELECTED_TRAITS)})
    con.register("selected_trait_names", selected_trait_frame)
    stage1_counts = con.execute(
        "SELECT 'GID' || cast(resolved_gid AS VARCHAR) AS canonical_gid, count(*) AS all_rows, "
        "sum(CASE WHEN accepted_canonical_trait IN (SELECT accepted_canonical_trait FROM selected_trait_names) THEN 1 ELSE 0 END) AS selected_rows "
        "FROM stage1 GROUP BY 1"
    ).fetch_df()
    count_map = stage1_counts.set_index("canonical_gid").to_dict("index")
    selected_gids = {gid for gid, counts in count_map.items() if int(counts["selected_rows"]) > 0}

    def state_summary(panel_id: str, gids: set[str], frame: pd.DataFrame, selected: bool) -> dict[str, object]:
        population = selected_gids if selected else stage1_gids
        linked = gids & population
        row_name = "selected_rows" if selected else "all_rows"
        return {
            "panel_id": panel_id,
            "population": "seven_selected_traits" if selected else "all_traits",
            "discovered_panel_samples": len(frame),
            "metadata_or_accepted_gid_count_all_panel": len(gids),
            "accepted_stage1_gids": len(linked),
            "stage1_population_gids": len(population),
            "stage1_rows_linked": int(sum(int(count_map[gid][row_name]) for gid in linked)),
            "stage1_population_rows": int(sum(int(counts[row_name]) for gid, counts in count_map.items())),
        }

    all_summaries: list[dict[str, object]] = []
    selected_summaries: list[dict[str, object]] = []
    for panel_id, group in samples.groupby("panel_id", sort=True):
        accepted = panel_gid_sets[panel_id]
        membership = membership_gid_sets[panel_id]
        all_row = state_summary(panel_id, accepted, group, False)
        selected_row = state_summary(panel_id, accepted, group, True)
        for selected, target in ((False, all_row), (True, selected_row)):
            population = selected_gids if selected else stage1_gids
            row_name = "selected_rows" if selected else "all_rows"
            membership_linked = membership & population
            target.update(
                {
                    "accepted_gid_count_all_panel": len(accepted),
                    "metadata_membership_gid_count_all_panel": len(membership),
                    "metadata_membership_stage1_gids": len(membership_linked),
                    "metadata_membership_stage1_rows": int(sum(int(count_map[gid][row_name]) for gid in membership_linked)),
                    "marker_present_gid_count_all_panel": len(marker_gid_sets[panel_id]),
                    "existing_qc_gid_count_all_panel": len(qc_gid_sets[panel_id]),
                    "strict_kernel_ready_gid_count_all_panel": len(strict_gid_sets[panel_id]),
                }
            )
        all_summaries.append(all_row)
        selected_summaries.append(selected_row)

    all_union = set().union(*panel_gid_sets.values()) if panel_gid_sets else set()
    metadata_union = set().union(*membership_gid_sets.values()) if membership_gid_sets else set()
    marker_union = set().union(*marker_gid_sets.values()) if marker_gid_sets else set()
    qc_union = set().union(*qc_gid_sets.values()) if qc_gid_sets else set()
    strict_union = set().union(*strict_gid_sets.values()) if strict_gid_sets else set()
    for selected, target in ((False, all_summaries), (True, selected_summaries)):
        summary = state_summary("ALL_PANEL_ACCEPTED_UNION", all_union, samples, selected)
        summary.update(
            {
                "accepted_gid_count_all_panel": len(all_union),
                "metadata_membership_gid_count_all_panel": len(metadata_union),
                "metadata_membership_stage1_gids": len(metadata_union & (selected_gids if selected else stage1_gids)),
                "metadata_membership_stage1_rows": int(sum(int(count_map[gid]["selected_rows" if selected else "all_rows"]) for gid in metadata_union & (selected_gids if selected else stage1_gids))),
                "marker_present_gid_count_all_panel": len(marker_union),
                "existing_qc_gid_count_all_panel": len(qc_union),
                "strict_kernel_ready_gid_count_all_panel": len(strict_union),
            }
        )
        target.append(summary)
    all_summary = pd.DataFrame(all_summaries)
    selected_summary = pd.DataFrame(selected_summaries)
    all_summary.to_csv(out_dir / "stage1_all_trait_linkage_summary.tsv", sep="\t", index=False)
    selected_summary.to_csv(out_dir / "stage1_selected_trait_linkage_summary.tsv", sep="\t", index=False)

    hmp = membership_gid_sets.get("frozen_hmp_v1", set()) & stage1_gids
    dartag = membership_gid_sets.get("dartag_panel2", set()) & stage1_gids
    hibap = membership_gid_sets.get("hibap35k", set()) & stage1_gids
    dh = dartag | hibap
    original_union = hmp | dh
    observed = {
        "stage1_gids": len(stage1_gids),
        "hmp": len(hmp),
        "dartag": len(dartag),
        "hibap": len(hibap),
        "dartag_or_hibap": len(dh),
        "hmp_and_dartag_or_hibap": len(hmp & dh),
        "original_union": len(original_union),
        "outside_original_union": len(stage1_gids - original_union),
        "selected_gids": len(selected_gids),
        "selected_hmp": len(hmp & selected_gids),
        "selected_original_union": len(original_union & selected_gids),
        "selected_rows": int(sum(int(counts["selected_rows"]) for counts in count_map.values())),
        "selected_rows_original_union": int(sum(int(count_map[gid]["selected_rows"]) for gid in original_union & selected_gids)),
    }
    reconciliation = pd.DataFrame(
        [
            {
                "metric": metric,
                "expected": expected,
                "observed": observed[metric],
                "difference": observed[metric] - expected,
                "status": "PASS" if observed[metric] == expected else "FAIL",
            }
            for metric, expected in EXPECTED_ORIGINAL.items()
        ]
    )
    reconciliation.to_csv(out_dir / "original_linkage_count_reconciliation.tsv", sep="\t", index=False)

    panel_membership = pd.DataFrame(
        [(panel, gid, "ACCEPTED_SAMPLE_LINK") for panel, gids in panel_gid_sets.items() for gid in gids]
        + [(panel, gid, "DOCUMENTED_METADATA_MEMBERSHIP") for panel, gids in membership_gid_sets.items() for gid in gids],
        columns=["panel_id", "canonical_gid", "linkage_state"],
    )
    con.register("panel_membership", panel_membership)
    per_trait = con.execute(
        """
        SELECT s.accepted_canonical_trait, p.panel_id, p.linkage_state,
               count(*) AS stage1_rows_linked,
               count(DISTINCT 'GID' || cast(s.resolved_gid AS VARCHAR)) AS stage1_gids_linked
        FROM stage1 s
        JOIN panel_membership p ON p.canonical_gid = 'GID' || cast(s.resolved_gid AS VARCHAR)
        GROUP BY 1,2,3 ORDER BY 1,2,3
        """
    ).fetch_df()
    per_trait.to_csv(out_dir / "stage1_linkage_by_trait_and_panel.tsv", sep="\t", index=False)
    union_frame = pd.DataFrame({"canonical_gid": sorted(all_union)})
    con.register("all_union", union_frame)
    per_trial = con.execute(
        """
        SELECT canonical_trial_name, count(*) AS stage1_rows,
               count(DISTINCT 'GID' || cast(resolved_gid AS VARCHAR)) AS stage1_gids,
               count(*) FILTER (WHERE u.canonical_gid IS NOT NULL) AS all_panel_linked_rows,
               count(DISTINCT u.canonical_gid) AS all_panel_linked_gids
        FROM stage1 s LEFT JOIN all_union u ON u.canonical_gid='GID'||cast(s.resolved_gid AS VARCHAR)
        GROUP BY 1 ORDER BY 1
        """
    ).fetch_df()
    per_trial.to_csv(out_dir / "stage1_linkage_by_trial.tsv", sep="\t", index=False)
    per_trial_cycle = con.execute(
        """
        SELECT canonical_trial_name, cycle, count(*) AS stage1_rows,
               count(DISTINCT 'GID' || cast(resolved_gid AS VARCHAR)) AS stage1_gids,
               count(*) FILTER (WHERE u.canonical_gid IS NOT NULL) AS all_panel_linked_rows,
               count(DISTINCT u.canonical_gid) AS all_panel_linked_gids
        FROM stage1 s LEFT JOIN all_union u ON u.canonical_gid='GID'||cast(s.resolved_gid AS VARCHAR)
        GROUP BY 1,2 ORDER BY 1,2
        """
    ).fetch_df()
    per_trial_cycle.to_csv(out_dir / "stage1_linkage_by_trial_cycle.tsv", sep="\t", index=False)

    overlap_rows = []
    panels = sorted(panel_gid_sets)
    for left in panels:
        for right in panels:
            overlap_rows.append(
                {
                    "panel_a": left,
                    "panel_b": right,
                    "accepted_gid_overlap_all_panel": len(panel_gid_sets[left] & panel_gid_sets[right]),
                    "stage1_gid_overlap": len(panel_gid_sets[left] & panel_gid_sets[right] & stage1_gids),
                }
            )
    pd.DataFrame(overlap_rows).to_csv(out_dir / "cross_panel_gid_overlap.tsv", sep="\t", index=False)
    con.close()
    return {
        "observed_original": observed,
        "original_reconciliation_pass": bool(reconciliation["status"].eq("PASS").all()),
        "stage1_gids": len(stage1_gids),
        "selected_gids": len(selected_gids),
        "all_panel_union_gids_total": len(all_union),
        "all_panel_union_stage1_gids": len(all_union & stage1_gids),
        "all_panel_union_selected_gids": len(all_union & selected_gids),
        "all_panel_union_stage1_rows": int(sum(int(count_map[gid]["all_rows"]) for gid in all_union & stage1_gids)),
        "all_panel_union_selected_rows": int(sum(int(count_map[gid]["selected_rows"]) for gid in all_union & selected_gids)),
        "all_panel_metadata_union_gids_total": len(metadata_union),
        "all_panel_metadata_union_stage1_gids": len(metadata_union & stage1_gids),
        "all_panel_metadata_union_stage1_rows": int(sum(int(count_map[gid]["all_rows"]) for gid in metadata_union & stage1_gids)),
        "all_panel_metadata_union_selected_gids": len(metadata_union & selected_gids),
        "all_panel_metadata_union_selected_rows": int(sum(int(count_map[gid]["selected_rows"]) for gid in metadata_union & selected_gids)),
        "strict_kernel_ready_panels": sorted(panel for panel, gids in strict_gid_sets.items() if gids),
    }


def glis_provenance(doi_ledger_path: Path, resolver_path: Path, phase3_root: Path, out_dir: Path) -> dict[str, object]:
    doi = pq.read_table(doi_ledger_path).to_pandas().fillna("")
    doi["DOI"] = doi["DOI"].astype(str).str.strip()
    valid = doi[doi["DOI"].str.fullmatch(r"10\.\d{4,9}/\S+", case=False, na=False)].copy()
    local = (
        valid.groupby("DOI", sort=True)
        .agg(
            local_record_rows=("DOI", "size"),
            local_source_files=("doi_source_file", lambda values: ";".join(sorted(set(map(str, values))))),
        )
        .reset_index()
    )
    resolver = pd.read_csv(resolver_path, sep="\t", dtype=str, keep_default_na=False)
    if resolver["DOI"].duplicated().any():
        raise RuntimeError("GLIS resolver DOI is not unique")
    linked = local.merge(resolver, on="DOI", how="left", validate="1:1", indicator=True)
    linked["typed_gid"] = linked["glis_gid"].map(lambda value: authoritative_gid(value, "glis_other_gid_field")[0])
    linked["identifier_semantics_status"] = np.where(
        linked["typed_gid"].ne(""),
        "PASS_GID_FROM_OFFICIAL_OTHER_GID_FIELD",
        "FAIL_NO_TYPED_OFFICIAL_GID",
    )
    linked["doi_digits_used_as_gid"] = False
    linked.to_parquet(out_dir / "glis_local_doi_gid_provenance.parquet", index=False)
    response_path = phase3_root / "glis_resolver_v2" / "glis_response_ledger.tsv"
    responses = pd.read_csv(response_path, sep="\t", dtype=str, keep_default_na=False)
    cache_failures = 0
    for row in responses.itertuples(index=False):
        cache_path = phase3_root / "glis_resolver_v2" / row.cache_file
        if not cache_path.exists() or sha256(cache_path) != row.response_sha256:
            cache_failures += 1
    summary = {
        "local_valid_doi_rows": len(valid),
        "local_unique_valid_dois": len(local),
        "local_dois_resolved": int(linked["typed_gid"].ne("").sum()),
        "local_dois_unresolved": int(linked["typed_gid"].eq("").sum()),
        "phase3_live_response_rows": len(responses),
        "phase3_live_single_gid_accepts": int(responses["parser_status"].eq("ACCEPT_EXACT_PAGE_DOI_SINGLE_GID").sum()),
        "phase3_response_cache_hash_failures": cache_failures,
        "doi_digits_used_as_gid": 0,
        "official_parser_rule": "matching requested DOI plus exactly one page-level GID <integer> token",
    }
    (out_dir / "glis_provenance_summary.json").write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def unresolved_phenotype_candidates(unresolved_path: Path, samples: pd.DataFrame, out_dir: Path) -> dict[str, int]:
    unresolved = pd.read_csv(unresolved_path, sep="\t", dtype=str, keep_default_na=False)
    evidence: dict[str, set[str]] = defaultdict(set)
    evidence_samples: dict[str, set[str]] = defaultdict(set)
    for row in samples[samples["accepted_canonical_gid"].ne("")].itertuples(index=False):
        values = [row.raw_sample_id]
        try:
            metadata_groups = json.loads(row.raw_metadata_json)
        except json.JSONDecodeError:
            metadata_groups = []
        for group in metadata_groups:
            values.extend(group.values())
        for value in values:
            normalized = re.sub(r"\s+", " ", clean(value)).upper()
            if normalized:
                evidence[normalized].add(row.accepted_canonical_gid)
                evidence_samples[normalized].add(row.panel_sample_key)
    generic = {
        "UNKNOWN", "DESCONOCIDO", "LOCAL CHECK", "LOCAL CHECK 1", "LOCAL CHECK 2",
        "CHECK", "CHECK 1", "CHECK 2", "TESTIGO", "TESTIGO LOCAL", "ENTRY", "LINE",
        "N/A", "NA", "NONE", "NULL", "0", "-", ".",
    }
    candidate_gids: list[str] = []
    candidate_samples: list[str] = []
    statuses: list[str] = []
    for value in unresolved["genotype_name"]:
        normalized = re.sub(r"\s+", " ", clean(value)).upper()
        gids = evidence.get(normalized, set())
        if not normalized or normalized in generic:
            statuses.append("NO_CANDIDATE_GENERIC_OR_BLANK_NAME")
            candidate_gids.append("")
            candidate_samples.append("")
        elif len(gids) == 1:
            statuses.append("CANDIDATE_REQUIRES_REVIEW_EXACT_TYPED_PANEL_METADATA_NAME")
            candidate_gids.append(next(iter(gids)))
            candidate_samples.append(";".join(sorted(evidence_samples[normalized])))
        elif len(gids) > 1:
            statuses.append("AMBIGUOUS_MULTIPLE_GIDS_FROM_EXACT_PANEL_METADATA_NAME")
            candidate_gids.append(";".join(sorted(gids)))
            candidate_samples.append(";".join(sorted(evidence_samples[normalized])))
        else:
            statuses.append("NO_CANONICAL_MATCH")
            candidate_gids.append("")
            candidate_samples.append("")
    unresolved["candidate_canonical_gids"] = candidate_gids
    unresolved["candidate_panel_sample_keys"] = candidate_samples
    unresolved["phase3g_review_status"] = statuses
    unresolved["applied_to_stage1"] = False
    unresolved["candidate_evidence_rule"] = np.where(
        unresolved["candidate_canonical_gids"].ne(""),
        "exact full-string match to typed panel metadata; candidate only",
        "none",
    )
    unresolved.to_parquet(out_dir / "unresolved_phenotype_identity_candidates.parquet", index=False)
    unresolved.to_csv(out_dir / "unresolved_phenotype_identity_candidates.tsv", sep="\t", index=False)
    return {
        "unresolved_keys_retained": len(unresolved),
        "unique_candidate_keys": int(pd.Series(statuses).str.startswith("CANDIDATE").sum()),
        "ambiguous_candidate_keys": int(pd.Series(statuses).str.startswith("AMBIGUOUS").sum()),
    }


def panel_inventory(files: pd.DataFrame, samples: pd.DataFrame, hmp_order_path: Path) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    panels = sorted(set(files["panel_id"]) | set(samples["panel_id"]))
    for panel in panels:
        panel_files = files[files["panel_id"].eq(panel)]
        panel_samples = samples[samples["panel_id"].eq(panel)]
        accepted = panel_samples[panel_samples["accepted_canonical_gid"].ne("")]
        rows.append(
            {
                "panel_id": panel,
                "platform": ";".join(sorted(set(panel_files["platform"]) - {""})) if not panel_files.empty else "frozen HMP kernel",
                "technology": ";".join(sorted(set(panel_files["technology"]) - {""})) if not panel_files.empty else "precomputed certified genomic kernel",
                "raw_file_count": len(panel_files),
                "raw_sample_count": len(panel_samples),
                "accepted_sample_count": len(accepted),
                "accepted_canonical_gid_count": accepted["accepted_canonical_gid"].nunique(),
                "marker_present_sample_count": int(panel_samples["marker_vector_present"].sum()),
                "existing_qc_pass_sample_count": int(panel_samples["existing_qc_status"].str.contains("PASS", na=False).sum()),
                "strict_kernel_ready_sample_count": int(panel_samples["kernel_readiness_status"].eq("STRICT_KERNEL_READY_EXISTING_QC").sum()),
                "manifest_available": bool(panel_samples["metadata_present"].any()),
                "existing_qc_documentation": ";".join(sorted(set(panel_samples["existing_qc_status"]) - {"NOT_ESTABLISHED"})),
                "parse_status": "PASS_PANEL_ACCOUNTED",
                "identity_status": "IDENTITIES_PARTIAL_OR_UNRESOLVED" if (panel_samples["accepted_canonical_gid"].eq("").any()) else "ALL_DISCOVERED_SAMPLES_TYPED",
                "source_files": ";".join(sorted(panel_files["relative_path"])),
            }
        )
    return pd.DataFrame(rows)


def validate_outputs(files: pd.DataFrame, samples: pd.DataFrame, linkage: dict[str, object], glis: dict[str, object]) -> pd.DataFrame:
    checks = [
        ("genotype_file_count_is_92", len(files) == 92, len(files)),
        ("every_genotype_file_has_terminal_disposition", files["terminal_inventory_disposition"].ne("").all(), int(files["terminal_inventory_disposition"].eq("").sum())),
        ("every_discovered_sample_has_terminal_mapping", samples["mapping_status"].ne("").all(), int(samples["mapping_status"].eq("").sum())),
        ("panel_sample_keys_unique", not samples["panel_sample_key"].duplicated().any(), int(samples["panel_sample_key"].duplicated().sum())),
        ("one_sample_at_most_one_accepted_gid", not samples["accepted_canonical_gid"].str.contains(";", regex=False).any(), int(samples["accepted_canonical_gid"].str.contains(";", regex=False).sum())),
        ("accepted_links_have_source_provenance", samples.loc[samples["accepted_canonical_gid"].ne(""), "source_locations"].ne("").all(), int(samples.loc[samples["accepted_canonical_gid"].ne(""), "source_locations"].eq("").sum())),
        ("no_accepted_cross_panel_candidate_only_link", not samples.loc[samples["accepted_canonical_gid"].ne(""), "evidence_tier"].str.contains("candidate_cross_panel", na=False).any(), int(samples.loc[samples["accepted_canonical_gid"].ne(""), "evidence_tier"].str.contains("candidate_cross_panel", na=False).sum())),
        ("original_linkage_counts_reproduced", linkage["original_reconciliation_pass"], linkage["observed_original"]),
        ("local_9072_glis_dois_resolved", glis["local_unique_valid_dois"] == 9072 and glis["local_dois_resolved"] == 9072, glis["local_dois_resolved"]),
        ("phase3_490_glis_responses_hash_valid", glis["phase3_live_response_rows"] == 490 and glis["phase3_response_cache_hash_failures"] == 0, {"rows": glis["phase3_live_response_rows"], "hash_failures": glis["phase3_response_cache_hash_failures"]}),
        ("doi_digits_never_used_as_gid", glis["doi_digits_used_as_gid"] == 0, glis["doi_digits_used_as_gid"]),
    ]
    return pd.DataFrame(
        [{"check": name, "status": "PASS" if passed else "FAIL", "observed": json.dumps(observed, sort_keys=True) if isinstance(observed, (dict, list)) else observed} for name, passed, observed in checks]
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--genotype-root", type=Path, default=Path("GENOTYPIC_DATA"))
    parser.add_argument("--raw-inventory", type=Path, required=True)
    parser.add_argument("--prior-profile", type=Path, default=Path("audit/genotypic_data_inventory.csv"))
    parser.add_argument("--phase3-root", type=Path, required=True)
    parser.add_argument("--stage1", type=Path, required=True)
    parser.add_argument("--hmp-order", type=Path, required=True)
    parser.add_argument("--doi-ledger", type=Path, required=True)
    parser.add_argument("--result-dir", type=Path, required=True)
    args = parser.parse_args()
    genotype_root = args.genotype_root.resolve()
    phase3_root = args.phase3_root.resolve()
    out_dir = args.result_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    files = inventory_files(args.raw_inventory.resolve(), args.prior_profile.resolve(), genotype_root)
    ledger = SampleLedger()
    add_hmp_reference(ledger, args.hmp_order.resolve())
    add_dartag(ledger, genotype_root)
    add_hibap(ledger, genotype_root)
    add_simple_header_gbs(ledger, genotype_root)
    seeds_dir = genotype_root / "Seeds_of_Discovery_-_MasAgro_Biodiversidad_Wheat_DArTseq-Derived_SNP_Data_Beta_Recall_Results_From_2011-2014"
    seeds_map = add_sample_crosswalk_panel(
        ledger,
        panel_id="seeds_of_discovery_dartseq",
        crosswalk_path=seeds_dir / "SampleIDvsGID_45610samples.txt",
        sample_column="SampleID",
        gid_column="GID",
        matrix_path=seeds_dir / "SEQ_SNPs_Extract_45610samples_102474markers.txt",
    )
    mexican_dir = genotype_root / "DArTseq-derived_SNPs_for_wheat_Mexican_landrace_accessions"
    mexican_map = add_sample_crosswalk_panel(
        ledger,
        panel_id="mexican_landrace_dartseq",
        crosswalk_path=mexican_dir / "Mexican_landrace_samples_for_Germinate.txt",
        sample_column="SampleID",
        gid_column="GID",
        matrix_path=mexican_dir / "SEQ_SNPs_Extract_mexican_8584samples_102474markers.txt",
    )
    add_mexican_results_gid_rows(ledger, genotype_root)
    add_80k_panels(
        ledger,
        genotype_root,
        [
            (normalize_path(seeds_dir / "SampleIDvsGID_45610samples.txt"), seeds_map),
            (normalize_path(mexican_dir / "Mexican_landrace_samples_for_Germinate.txt"), mexican_map),
        ],
    )
    add_cimmyt_bread_panel(ledger, genotype_root)
    add_haplotype_panel(ledger, genotype_root)
    add_mas_xlsx_panels(ledger, genotype_root)
    add_45ibwsn_mas(ledger, genotype_root)

    samples = ledger.finalize()
    if samples.empty:
        raise RuntimeError("No panel samples discovered")
    samples.to_parquet(out_dir / "sample_identifier_ledger.parquet", index=False)
    crosswalk_columns = [
        "panel_id", "panel_sample_key", "raw_sample_id", "accepted_canonical_gid",
        "mapping_status", "evidence_tier", "evidence_type", "ambiguity_set",
        "conflict_status", "source_files", "source_locations", "normalization_rules",
    ]
    samples[crosswalk_columns].to_parquet(out_dir / "sample_gid_crosswalk.parquet", index=False)
    evidence = pd.DataFrame(ledger.evidence)
    evidence.to_parquet(out_dir / "linkage_evidence_ledger.parquet", index=False)
    unresolved_samples = samples[~samples["mapping_status"].str.startswith("ACCEPTED")].copy()
    unresolved_samples.to_csv(out_dir / "unmatched_ambiguous_conflicting_samples.tsv", sep="\t", index=False)
    marker_columns = [
        "panel_id", "panel_sample_key", "raw_sample_id", "accepted_canonical_gid",
        "metadata_present", "marker_vector_present", "existing_imputed_matrix_present",
        "existing_kernel_order_present", "existing_qc_status", "audit_qc_status",
        "kernel_readiness_status", "exclusion_or_unresolved_reason", "matrix_order_first",
    ]
    samples[marker_columns].to_parquet(out_dir / "marker_presence_and_qc.parquet", index=False)
    collisions = build_namespace_collisions(samples)
    collisions.to_csv(out_dir / "namespace_collision_ledger.tsv", sep="\t", index=False)
    collisions.to_parquet(out_dir / "namespace_collision_ledger.parquet", index=False)
    duplicates = build_duplicate_report(samples)
    duplicates.to_csv(out_dir / "cross_panel_duplicate_report.tsv", sep="\t", index=False)
    orders = write_orders(samples, out_dir)
    orders.to_csv(out_dir / "sample_order_manifest.tsv", sep="\t", index=False)

    files = profile_genotype_files(files, samples)
    files.to_csv(out_dir / "genotype_file_inventory.tsv", sep="\t", index=False)
    panels = panel_inventory(files, samples, args.hmp_order.resolve())
    panels.to_csv(out_dir / "panel_inventory.tsv", sep="\t", index=False)

    linkage = stage1_linkage(args.stage1.resolve(), samples, evidence, out_dir)
    unresolved_summary = unresolved_phenotype_candidates(
        phase3_root / "gid_coverage_release_v1" / "unresolved_numeric_identity_keys.tsv",
        samples,
        out_dir,
    )
    glis = glis_provenance(
        args.doi_ledger.resolve(),
        phase3_root / "glis_resolver_v2" / "glis_resolver_v2.tsv",
        phase3_root,
        out_dir,
    )
    validation = validate_outputs(files, samples, linkage, glis)
    validation.to_csv(out_dir / "validation_checks_stage1.tsv", sep="\t", index=False)
    summary = {
        "status": "PASS_STAGE1_PHASE3G_AUDIT_SCRIPT" if validation["status"].eq("PASS").all() else "FAIL_STAGE1_PHASE3G_AUDIT_SCRIPT",
        "version": VERSION,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "genotype_files": len(files),
        "panels": len(panels),
        "panel_samples": len(samples),
        "accepted_panel_samples": int(samples["accepted_canonical_gid"].ne("").sum()),
        "unique_accepted_gids_all_panels": int(samples.loc[samples["accepted_canonical_gid"].ne(""), "accepted_canonical_gid"].nunique()),
        "marker_present_samples": int(samples["marker_vector_present"].sum()),
        "existing_qc_pass_samples": int(samples["existing_qc_status"].str.contains("PASS", na=False).sum()),
        "strict_kernel_ready_samples": int(samples["kernel_readiness_status"].eq("STRICT_KERNEL_READY_EXISTING_QC").sum()),
        "unmatched_or_ambiguous_samples": len(unresolved_samples),
        "namespace_collision_rows": len(collisions),
        "duplicate_gid_rows": len(duplicates),
        "linkage": linkage,
        "glis": glis,
        "unresolved_phenotypes": unresolved_summary,
        "validation_passed": int(validation["status"].eq("PASS").sum()),
        "validation_total": len(validation),
    }
    (out_dir / "phase3g_audit_summary.json").write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
